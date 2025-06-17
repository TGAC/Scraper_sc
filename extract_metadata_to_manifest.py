import os

import pandas as pd
import fitz  # PyMuPDF
from openai import OpenAI
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import json
from pathlib import Path
import shutil

# --- Config ---
excel_file = "sc_rnaseq_mixs_v0.1_base_unprotected.xlsx"

# --- OpenAI Client ---
client = OpenAI(
    api_key=os.getenv("GPT_KEY"))


# --- PDF text extraction ---
def extract_pdf_text(pdf_path):
    doc = fitz.open(pdf_path)
    return "\n".join([page.get_text() for page in doc])


# --- Prepare base conversation context ---
def get_base_messages(pdf_text):
    return [
        {
            "role": "system",
            "content": (
                "You are an expert in single cell genomics. "
                "Here are some topics you are well versed in:"
                "Single-cell genomics primarily involves isolating individual cells, analyzing their genetic material (DNA, RNA, or epigenome), and processing the data to understand cell-to-cell differences and relationships. This technology allows researchers to study the unique molecular profiles of individual cells, providing deeper insights into biological processes and diseases. "
                "Here's a more detailed breakdown of the main aspects:"
                "1. Cell Isolation and Preparation:"
                "Single-cell isolation: The process of separating individual cells from a cell population. "
                "Genetic material extraction and amplification: Extracting and amplifying the DNA or RNA from each isolated cell."
                "Library preparation: Preparing the genetic material for sequencing. "
                "2. Sequencing and Analysis:"
                "Sequencing: Using next-generation sequencing technologies to analyze the DNA or RNA. "
                "Data analysis: Analyzing the sequencing data to understand cell-to-cell differences in gene expression, DNA variations, or epigenetic modifications. "
                "3. Key Applications and Insights:"
                "Cell type identification and classification: Identifying and classifying different cell types based on their unique molecular profiles. "
                "Understanding cell states and heterogeneity: Exploring the diversity of cellular states within a population and identifying rare cell types."
                "Investigating biological processes: Understanding how cells interact with each other and their environment, and how these interactions affect development, disease, and other biological processes. "
                "Identifying genetic variations: Detecting genetic mutations or variations within individual cells. "
                "Analyzing epigenetic modifications: Studying how modifications to DNA can affect gene expression and cellular behavior. "
                "Developing new diagnostics and therapeutics: Identifying biomarkers and drug targets based on single-cell data. "
                "4. Types of Single-Cell Genomics:"
                "Single-cell RNA sequencing (scRNA-seq):"
                "Analyzing RNA transcripts in individual cells to understand gene expression patterns. "
                "Single-cell DNA sequencing."
                "Sequencing the entire genome or specific regions of interest in individual cells. "
                "Single-cell epigenomics:"
                "Studying epigenetic modifications, such as DNA methylation and histone modifications, in individual cells. "
                "5. Advantages of Single-Cell Genomics:"
                "Higher resolution: Provides a more detailed view of cellular heterogeneity and cell-to-cell differences. "
                "Improved understanding of biological processes: Reveals how cells interact with each other and their environment. "
                "Potential for new diagnostics and therapeutics: Identifies biomarkers and drug targets based on single-cell data. "

                "You are extracting metadata from a scientific paper "
                "to fill out fields in several Excel worksheets, one per metadata type. Maintain consistent, structured answers. "
                "Where multiple elements exist (e.g. multiple people or samples), return one item per row. Never combine items into arrays. "
                "Do not guess — leave missing values blank. Create unique ids for the '*_id' fields and maintain their consistency across the entire workbook."
                "Fields which are listed as optional can be left blank, however, fields which are not, you must try as hard as possible, to find an accurate value for."
            )
        },
        {
            "role": "user",
            "content": f"Here is the text of the paper:\n\n{pdf_text}"
        }
    ]


# --- GPT query ---
def query_gpt_for_sheet(messages, sheet_name, fields):
    field_prompt = (
        f"Extract the following metadata fields for the '{sheet_name}' worksheet. "
        f"Return data as JSON. Fields:\n{fields}"
    )
    messages.append({"role": "user", "content": field_prompt})
    '''
    response = client.chat.completions.create(
        model="gpt4.1",
        messages=messages,
        temperature=0,
        max_tokens=32768
    )
    '''
    response = client.chat.completions.create(
        model="o3",
        messages=messages,
        temperature=1,
        max_completion_tokens=32768
    )
    reply = response.choices[0].message.content.strip()
    messages.append({"role": "assistant", "content": reply})
    return reply


# --- Main process ---
input_pdf_path = Path("pdfs")
for file in input_pdf_path.iterdir():
    if file.is_file():
        if ".DS_Store" in file.name:
            continue
        updated_file = Path("completed_manifests/" + str(file.name)).with_suffix(".xlsx")
        pdf_file = file
        pdf_text = extract_pdf_text(pdf_file)

        xls = pd.ExcelFile(excel_file)
        # sheets_to_process = ["study"]
        sheets_to_process = ["study", "person", "sample", "dissociation", "cell_suspension", "lib_prep", "sequencing"]

        # Cache dataframes for column width adjustment
        generated_dfs = {}

        # Create GPT session context
        conversation = get_base_messages(pdf_text)

        # --- Write updated data to Excel ---
        with pd.ExcelWriter(updated_file, engine='openpyxl') as writer:
            for sheet in sheets_to_process:
                if sheet not in xls.sheet_names:
                    continue

                df_headers = pd.read_excel(xls, sheet_name=sheet, engine='openpyxl', nrows=0)
                fields = df_headers.columns.tolist()

                print(f"🔍 Extracting metadata for sheet: {sheet}")
                gpt_response = query_gpt_for_sheet(conversation, sheet, fields)

                try:
                    metadata = json.loads(gpt_response)
                except json.JSONDecodeError:
                    print(f"⚠️ Warning: Failed to parse GPT response for sheet '{sheet}'. Response:\n{gpt_response}")
                    metadata = [{col: "Not Provided" for col in fields}]

                if isinstance(metadata, list):
                    new_df = pd.DataFrame(metadata)
                else:
                    new_df = pd.DataFrame([{col: metadata.get(col, "Not Provided") for col in fields}])

                new_df.to_excel(writer, sheet_name=sheet, index=False)
                generated_dfs[sheet] = new_df

            if not generated_dfs:
                raise RuntimeError("❌ No sheets were processed. Check API key or GPT responses.")

        # print(f"✅ Metadata inserted and saved to {updated_file}")

        # --- Adjust column widths and ensure sheets are visible ---
        wb = load_workbook(updated_file)

        for sheet_name, df in generated_dfs.items():
            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]
            ws.sheet_state = "visible"

            for i, column in enumerate(df.columns, start=1):
                max_len = max(df[column].astype(str).map(len).max(), len(str(column)))
                col_letter = get_column_letter(i)
                ws.column_dimensions[col_letter].width = max_len + 2

        # --- Ensure at least one sheet is visible ---
        if not any(wb[s].sheet_state == "visible" for s in wb.sheetnames):
            wb[wb.sheetnames[0]].sheet_state = "visible"

        wb.save(updated_file)
        # Move the processed PDF to the 'done' folder
        done_path = Path("done") / file.name
        done_path.parent.mkdir(parents=True, exist_ok=True)  # Ensure 'done' directory exists
        shutil.move(str(file), str(done_path))
        print(f"moved {file.name} to {done_path}")
        print(f"🎉 All done! Updated file saved to: {updated_file}\n\n")
